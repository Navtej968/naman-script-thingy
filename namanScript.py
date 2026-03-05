import re
import asyncio
import aiohttp
import base64
from bs4 import BeautifulSoup
from urllib.parse import urlparse, urljoin
import pandas as pd
from openpyxl import load_workbook
from collections import deque
import sys


if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

INPUT_FILE = "work sheet 2.xlsx"
OUTPUT_FILE = "input_updated.xlsx"
START_ROW = 1   # change if needed (1 = first row)

MAX_PAGES_CRAWL = 5
CONCURRENT_REQUESTS = 50

EMAIL_REGEX = r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\b"

PAGES_TO_CHECK = [
    "", "/contact", "/contact-us", "/about", "/about-us",
    "/support", "/help", "/team", "/careers"
]

PRIORITY_KEYWORDS = ["info", "support", "contact", "hello", "sales", "business"]
IGNORE_KEYWORDS = ["noreply", "no-reply", "donotreply"]

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "en-US,en;q=0.9"
}


def get_domain(url):
    return urlparse(url).netloc.replace("www.", "").lower()

def clean_email(email):
    if not email:
        return ""
    email = email.lower().strip()
    email = re.sub(r"[^a-z0-9@._+-]", "", email)
    if email.count("@") != 1:
        return ""
    return email

def select_best_email(emails):
    emails = [e for e in emails if not any(k in e for k in IGNORE_KEYWORDS)]
    for key in PRIORITY_KEYWORDS:
        for e in emails:
            if e.startswith(key):
                return e
    return sorted(emails)[0] if emails else ""

def is_js_only(html):
    soup = BeautifulSoup(html, "html.parser")
    return len(soup.get_text(strip=True)) < 200 and any(
        k in html.lower() for k in ["__next", "react", "angular", "vue"]
    )

def extract_emails(text):
    return set(re.findall(EMAIL_REGEX, text or "", re.I))

def extract_base64_emails(text):
    found = set()
    for chunk in re.findall(r'[A-Za-z0-9+/=]{40,200}', text):
        try:
            decoded = base64.b64decode(chunk).decode(errors="ignore")
            found.update(extract_emails(decoded))
        except:
            pass
    return found

def extract_emails_from_html(html):
    emails = set()
    soup = BeautifulSoup(html, "html.parser")

    # mailto links
    for a in soup.find_all("a", href=True):
        if a["href"].lower().startswith("mailto:"):
            emails.add(clean_email(a["href"][7:].split("?")[0]))

    # visible text
    text = soup.get_text(" ", strip=True)
    emails.update(clean_email(e) for e in extract_emails(text))

    # scripts
    for s in soup.find_all("script"):
        if s.string:
            emails.update(clean_email(e) for e in extract_emails(s.string))

    # base64
    emails.update(clean_email(e) for e in extract_base64_emails(html))

    return {e for e in emails if e}

sem = asyncio.Semaphore(CONCURRENT_REQUESTS)

async def fetch(session, url):
    async with sem:
        try:
            async with session.get(url, timeout=8) as r:
                if r.status == 200:
                    return await r.text()
        except:
            pass
    return ""

async def scrape_contact_pages(base_url, session):
    for path in PAGES_TO_CHECK:
        html = await fetch(session, urljoin(base_url, path))
        if not html:
            continue
        email = select_best_email(extract_emails_from_html(html))
        if email:
            return email
    return ""

async def crawl_domain_for_email(base_url, session):
    domain = get_domain(base_url)
    visited, queue = set(), deque([base_url])
    found = set()

    while queue and len(visited) < MAX_PAGES_CRAWL:
        url = queue.popleft()
        if url in visited:
            continue
        visited.add(url)

        html = await fetch(session, url)
        if not html:
            continue

        found.update(extract_emails_from_html(html))
        if found:
            return select_best_email(found)

        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            link = urljoin(base_url, a["href"])
            if get_domain(link) == domain:
                queue.append(link)

    return select_best_email(found)

async def fetch_email(row, url, session):
    html = await fetch(session, url)
    if html and not is_js_only(html):
        email = select_best_email(extract_emails_from_html(html))
        if email:
            return row, email

    email = await scrape_contact_pages(url, session)
    if email:
        return row, email

    email = await crawl_domain_for_email(url, session)
    if email:
        return row, email

    return row, "could not find"


async def main():
    wb = load_workbook(INPUT_FILE)
    ws = wb.active
    df = pd.read_excel(INPUT_FILE, header=None)

    targets = []
    for i in range(START_ROW - 1, len(df)):
        url = str(df.iloc[i, 0]).strip()
        if url.startswith("http"):
            targets.append((i, url))

    connector = aiohttp.TCPConnector(limit=200, limit_per_host=10, ssl=False)
    timeout = aiohttp.ClientTimeout(total=8)

    async with aiohttp.ClientSession(
        headers=HEADERS,
        connector=connector,
        timeout=timeout
    ) as session:

        tasks = [fetch_email(row, url, session) for row, url in targets]
        for i, task in enumerate(asyncio.as_completed(tasks), 1):
            row, email = await task
            ws.cell(row=row + 1, column=2, value=email)
            print(f"({i}/{len(tasks)}) Row {row + 1}: {email}")

    wb.save(OUTPUT_FILE)
    print(f"\n✅ Done! Saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    asyncio.run(main())
